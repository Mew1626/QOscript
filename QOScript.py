import os
import datetime
import sys
import csv

from tkinter import Tk
from tkinter import StringVar
from tkinter import N, W, E, S
from tkinter import filedialog
from tkinter import ttk

from openpyxl import load_workbook
from os.path import expanduser

import logging
logging.basicConfig(level='INFO', filename='QOScript.log', format='%(asctime)s :: %(levelname)s :: %(message)s')

warantycsv = ""
purchasecsv = ""
evalcsv = ""
loncsv = ""

exQO1 = ""
exQO2 = ""
exQO3 = ""
exQO4 = ""

#GUI Section
root = Tk()
reports = StringVar()
under10min = StringVar()
monthselection = 0

def wrap(): #runs the whole shebang
    logging.info("QO functions have started to run")
    try:
        QO1(exQO1, reports.get(), under10min.get(), month.current()+2)
    except:
        logging.error("QO1 Failed")
    try:
        QO2(exQO2, purchasecsv, warantycsv)
    except:
        logging.error("QO2 Failed")
    try:
        QO3(exQO3, evalcsv) 
    except:
        logging.error("QO3 Failed")
    try:
        QO4(exQO4, loncsv) 
    except:
        logging.error("QO4 Failed")

def browse(iden):
    global warantycsv, purchasecsv, evalcsv, loncsv
    home = expanduser("~")
    root.filename = filedialog.askopenfilename(initialdir = home,title = "Select file",filetypes = (("csv file","*.csv"),("all files","*.*")))
    if iden == 0: 
        purchasecsv = root.filename
        pur.set(root.filename)
    elif iden == 1: 
        warantycsv = root.filename
        war.set(root.filename)
    elif iden == 2: 
        evalcsv = root.filename
        eva.set(root.filename)
    elif iden == 3: 
        loncsv = root.filename
        lon.set(root.filename)

def QObrowse(iden):
    global exQO1, exQO2, exQO3, exQO4
    home = expanduser("~")
    root.filename = filedialog.askopenfilename(initialdir = home,title = "Select file",filetypes = (("excel file","*.xlsx"),("all files","*.*")))
    if iden == 0: 
        exQO1 = root.filename
        QS1.set(root.filename)
    elif iden == 1: 
        exQO2 = root.filename
        QS2.set(root.filename)
    elif iden == 2: 
        exQO3 = root.filename
        QS3.set(root.filename)
    elif iden == 3: 
        exQO4 = root.filename
        QS4.set(root.filename)

def gui():

    root.title("Quality Objectives")

    mainframe = ttk.Frame(root, padding="3 3 12 12")
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    #actual gui elements
    reports_entry = ttk.Entry(mainframe, width=7, textvariable=reports)
    under10min_entry = ttk.Entry(mainframe, width=7, textvariable=under10min)
    reports_entry.grid(column=0, row= 2, sticky= (W, E))
    under10min_entry.grid(column=1, row= 2, sticky= (W, E))

    ttk.Label(mainframe, text="Reports").grid(column=0, row= 1, sticky= (W, E))
    ttk.Label(mainframe, text="Under 10 Min").grid(column=1, row= 1, sticky= (W, E))

    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    ttk.Label(mainframe, text="Month:").grid(column=0, row= 3, sticky= (W, E))
    monthvar = StringVar()
    global month
    month = ttk.Combobox(mainframe, textvariable=monthvar)
    month['values'] = months
    month.grid(column=1, row= 3, sticky= (W, E))
    month.current(datetime.date.today().month - 2)

    ttk.Button(mainframe, text="Execute", command=wrap).grid(column=2, row= 13, sticky= (W, E))

    global pur, war, eva, lon 
    pur = StringVar()
    war = StringVar()
    eva = StringVar()
    lon = StringVar()

    ttk.Label(mainframe, text="Purchase CSV").grid(column=0, row= 4, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=pur).grid(column=1, row=4, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:browse(0)).grid(column=2, row= 4, sticky= (W, E))
    
    ttk.Label(mainframe, text="Waranty CSV").grid(column=0, row= 5, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=war).grid(column=1, row=5, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:browse(1)).grid(column=2, row= 5, sticky= (W, E))
    
    ttk.Label(mainframe, text="Eval CSV").grid(column=0, row= 6, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=eva).grid(column=1, row=6, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:browse(2)).grid(column=2, row= 6, sticky= (W, E))
    
    ttk.Label(mainframe, text="Lon CSV").grid(column=0, row= 7, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=lon).grid(column=1, row=7, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:browse(3)).grid(column=2, row= 7, sticky= (W, E))

    global QS1, QS2, QS3, QS4
    QS1 = StringVar()
    QS2 = StringVar()
    QS3 = StringVar()
    QS4 = StringVar()

    ttk.Label(mainframe, text="QO1 Excel File").grid(column=0, row= 9, sticky= (W, E))
    ttk.Entry(mainframe, textvariable= QS1).grid(column=1, row=9, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:QObrowse(0)).grid(column=2, row= 9, sticky= (W, E))
    
    ttk.Label(mainframe, text="QO2 Excel File").grid(column=0, row= 10, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=QS2).grid(column=1, row=10, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:QObrowse(1)).grid(column=2, row= 10, sticky= (W, E))
    
    ttk.Label(mainframe, text="QO3 Excel File").grid(column=0, row= 11, sticky= (W, E))
    ttk.Entry(mainframe, textvariable=QS3).grid(column=1, row=11, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:QObrowse(2)).grid(column=2, row= 11, sticky= (W, E))
    
    ttk.Label(mainframe, text="QO4 Excel File").grid(column=0, row= 12, sticky=(W, E))
    ttk.Entry(mainframe, textvariable=QS4).grid(column=1, row=12, sticky=(W, E))
    ttk.Button(mainframe, text="Browse", command=lambda:QObrowse(3)).grid(column=2, row= 12, sticky= (W, E))
    
    ttk.Label(mainframe, text="If anything breaks there should be a log file created in the same directory as this script.").grid( row=13, columnspan=2, sticky=(W, E))

    for child in mainframe.winfo_children(): 
        child.grid_configure(padx=5, pady=5)
    root.mainloop()

#QO1 
def QO1(excel_file, reports, under10min, month):
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    ws.cell(month,3,reports)
    ws.cell(month,4,under10min)
    logging.info("QO1 Done!")
    wb.save(excel_file)

#QO2
def QO2(excel_file, purchasecsv, warantycsv):
    wb = load_workbook(excel_file)

    #load Purchases sheet
    ws = wb.worksheets[1]
    wsname = wb.sheetnames[1]
 
    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 1)
    ws = wb.worksheets[1]
 
    #import the CSV data
    try:
        with open(purchasecsv) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
    except FileNotFoundError as fnf_error:
        logging.error(fnf_error)
    
    #load Waranty sheet
    ws = wb.worksheets[2]
    wsname = wb.sheetnames[2]
 
    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 2)
    ws = wb.worksheets[2]
    
    #import the CSV data
    try:
        with open(warantycsv) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
    except FileNotFoundError as fnf_error:
       logging.error(fnf_error)
    
    #load Overlap sheet
    ws = wb.worksheets[3]
    wsname = wb.sheetnames[3]

    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 3)
    ws = wb.worksheets[3]
    
    # load data from worksheets
    wsp = wb.worksheets[1]
    wsw = wb.worksheets[2]
    cmax = wsw.max_row

    #sanity variables
    punum = (int(wsp.max_row) - 1)
    pssn,wssn,returns,data = [],[],[],[]
    date = wsw.cell(2,1).value[1:-6]

    #transfer and transform
    #Creates rows for the overlap sheet
    for i in range(1, cmax+1):
        if i != 1:
            pssn.append(wsp.cell(i,4).value)
            wssn.append(wsw.cell(i,4).value)
        data.append([wsp.cell(i,1).value,wsp.cell(i,2).value,wsp.cell(i,4).value,wsp.cell(i,5).value,None,None,wsw.cell(i,1).value,wsw.cell(i,2).value,wsw.cell(i,4).value,wsw.cell(i,5).value,wsw.cell(i,6).value])
   
    #finds the returns and adds them to the "returns" array
    for i in range(len(pssn)):
        for j in range(len(pssn)):
            if bool(pssn[i] == wssn[j]) & bool(pssn[i] != None) & bool(pssn[i] != ''):
                returns.append(wsp.cell(i+2,2).value)
   
    #adds the rows to the overlap sheet
    for row in data:
        ws.append(row)
    
    #simply adds the precalculated returns 
    ws = wb.worksheets[4]
    for i in range(1, len(returns)+1):
        ws.cell(i+1,int(date),returns[i-1])
    
    #updates the oversite sheet 
    ws = wb.worksheets[0]
    ws.cell(int(date)+1,3,punum)
    ws.cell(int(date)+1,4,len(returns))
    logging.info("QO2 Done!")
    wb.save(excel_file)

#QO3
def QO3(excel_file, evalcsv):
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    rows = []
    global total, ontim3
    total = 0 
    ontim3 = 0

    try:
        with open(evalcsv) as f:#open's csv
            reader = csv.reader(f, delimiter=',')
            temp = getmaxrow(ws)
            for row in reader:
                if row[0] != "One Day Trial?": #skips the colum name row
                    rows.append([row[4], row[16], row[17], QO3ontime(row[16], row[17])])# gets the data for the Dates function
                else:
                     rows.append(["Account Name", row[16], row[17], "On Time?"])
                for i in range(len(row)):#moves csv data into excel sheet
                    if i >= 2 :
                        ws.cell(temp, i-1, row[i])
                temp += 1

    except FileNotFoundError as fnf_error:
        logging.error(fnf_error)
    
    #Dates
    ws = wb.worksheets[1]
    temp = getmaxrow(ws)
    for row in rows: #inputs the data made in Lon into the Relavant sheet
        for i in range(len(row)):
            ws.cell(temp, i+1, row[i])
        temp += 1

    #Count
    ws = wb.worksheets[2] # inputs the on time and total variables
    ws.cell(int(month)+1,2,ontim3) 
    ws.cell(int(month)+1,3,total)
    
    #Overview
    ws = wb.worksheets[3]
    ws.cell(int(month)+1,2,ontim3 / total) #inputs the percent

    logging.info("QO3 Done !")
    wb.save(excel_file)

def getmaxrow(ws):
    maybe = 0
    for i in range(1, ws.max_row):
       if (ws.cell(i,1).value == "" or ws.cell(i,1).value == None) and maybe == 1:
           return(i)
       elif ws.cell(i,1).value == "" or ws.cell(i,1).value == None:
           maybe = 1
       else:
           maybe = 0
       if i == (ws.max_row - 1):
           return (ws.max_row + 2)

def QO3ontime(ship, train): # figures out if the ship date is before training and install date
    global total, ontim3, month
    total = total + 1
    ship = ship.split("-")
    train = train.split("-")
    month = ship[0]
    if (ship[0] <= train[0] and ship[2] <= train[2]) or (ship[1] <= train[1] and ship[0] == train[0]):
        ontim3 = ontim3 + 1
        return("Yes")
    else: return("No")

#QO4
def QO4(excel_file, loncsv):
    global total, late
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    rows = []
    total, late = 0, 0 

    try:
        with open(loncsv) as f:#open's csv
            reader = csv.reader(f, delimiter=',')
            temp = getmaxrow(ws)
            for row in reader:
                if row[0] != "Purchase order Date": #skips the colum name row
                    rows.append([row[1], row[3], row[0], row[13], QO4ontime(row[0], row[13])])# gets the data for the Dates function
                else: 
                    rows.append([row[1], row[3], row[0], row[13], "Within 48 Hours?"])
                for i in range(len(row)):
                    ws.cell(temp, i+1, row[i])
                temp += 1
    except FileNotFoundError as fnf_error:
        logging.error(fnf_error)

    #Revalant
    ws = wb.worksheets[1]
    temp = getmaxrow(ws)
    for row in rows: #inputs the data made in Lon into the Relavant sheet
        for i in range(len(row)):
            ws.cell(temp, i+1, row[i])
        temp += 1
    
    #Overview
    ws = wb.worksheets[2]
    ws.cell(int(month)+1,2,total) 
    ws.cell(int(month)+1,3,late) 

    logging.info("QO4 Done !")
    wb.save(excel_file)
   
def QO4ontime(pur, ship): 
    global total, late, month
    total = total + 1
    if ship == "" or ship == None:
        return("")
    ship = ship.split("-")
    pur = pur.split("-")
    month = pur[0]
    if ((int(pur[1]) - int(ship[1])) >= -2) and pur[0] == ship[0]:
        return("Yes")
    elif datetime.date((2000+int(pur[2])),int(pur[0]),int(pur[1])).weekday() >= 4 and datetime.date((2000+int(ship[2])),int(ship[0]),int(ship[1])).weekday() <= 2 and pur[0] == ship[0] and pur[1] <= ship[1]: 
        #second condition is wednsday (i.e the 2) monday is 0 and saunday is 6
        return("Yes")
    else:
        late += 1
        return ("No")

if __name__ == '__main__':
    logging.info("Started")
    gui()