import csv
import datetime
import sys
from openpyxl import load_workbook

#These 2 are the only variables you need to change.(normally you can just pass them in through command line arguments i.e. "python QO4.py $excelfile $csv")
excel_file = sys.argv[1]
loncsv = sys.argv[2]

#no touchy the rest unless it is borked
#if borked the function names are what sheet in the excel file it changes 

def Lon():
    global total, late, wb, rows
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    rows = []
    total, late = 0, 0 

    try:
        with open(loncsv) as f:#open's csv
            reader = csv.reader(f, delimiter=',')
            temp = getmaxrow(ws)
            for row in reader:
                if row[0] != "Purchase order Date": #skips the colum name row
                    rows.append([row[1], row[3], row[0], row[13], ontime(row[0], row[13])])# gets the data for the Dates function
                else: 
                    rows.append([row[1], row[3], row[0], row[13], "Within 48 Hours?"])
                for i in range(len(row)):
                    ws.cell(temp, i+1, row[i])
                temp += 1
    except FileNotFoundError as fnf_error:
        print(fnf_error)

# figures out if the ship date is within 48 hours(takes into account weekends, basically if a purchase order comes in on or after friday they have until wednesday to send it)
def ontime(pur, ship): 
    global total, late, month
    total = total + 1
    if ship == "" or ship == None:
       return("")
    ship = ship.split("-")
    pur = pur.split("-")
    month = pur[0]
    if ((int(pur[1]) - int(ship[1])) >= -2) and pur[0] == ship[0]:
        return("Yes")
    elif datetime.date((2000+int(pur[2])),int(pur[0]),int(pur[1])).weekday() >= 4 and datetime.date((2000+int(ship[2])),int(ship[0]),int(ship[1])).weekday() <= 2 and pur[0] == ship[0] and pur[1] <= ship[1]: 
        #second condition is wednsday (i.e the 2) monday is 0 and saunday is 6
        return("Yes")
    else:
        late += 1
        return ("No")

def Relevant():
    ws = wb.worksheets[1]
    temp = getmaxrow(ws)
    for row in rows: #inputs the data made in Lon into the Relavant sheet
        for i in range(len(row)):
            ws.cell(temp, i+1, row[i])
        temp += 1

def Overview():
    ws = wb.worksheets[2]
    ws.cell(int(month)+1,2,total) 
    ws.cell(int(month)+1,3,late) 

def getmaxrow(ws):
     maybe = 0
     for i in range(1, ws.max_row):
        if (ws.cell(i,1).value == "" or ws.cell(i,1).value == None) and maybe == 1:
            return(i)
        elif ws.cell(i,1).value == "" or ws.cell(i,1).value == None:
            maybe = 1
        else:
            maybe = 0
        if i == (ws.max_row - 1):
            return (ws.max_row + 2)

def main():
    Lon()
    Relevant()
    Overview()
    print("QO4 Done !")
    wb.save(excel_file)

main()