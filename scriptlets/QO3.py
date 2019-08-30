import csv
import sys
from openpyxl import load_workbook

#These 2 are the only variables you need to change.(normally you can just pass them in through command line arguments i.e. "python QO3.py $excelfile $csv")

excel_file = sys.argv[1]
evalcsv = sys.argv[2]


#no touchy the rest unless it is borked
#if borked the function names are what sheet in the excel file it changes (most of the computation of shit is done in the Eval function)

def Eval():
    global total, ontim3, wb, rows
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    rows = []
    total = 0 
    ontim3 = 0

    try:
        with open(evalcsv) as f:#open's csv
            reader = csv.reader(f, delimiter=',')
            temp = getmaxrow(ws)
            for row in reader:
                if row[0] != "One Day Trial?": #skips the colum name row
                    rows.append([row[4], row[16], row[17], ontime(row[16], row[17])])# gets the data for the Dates function
                else:
                     rows.append(["Account Name", row[16], row[17], "On Time?"])
                for i in range(len(row)):#moves csv data into excel sheet
                    if i >= 2 :
                        ws.cell(temp, i-1, row[i])
                temp += 1

    except FileNotFoundError as fnf_error:
        print(fnf_error)

def Dates():
    ws = wb.worksheets[1]
    temp = getmaxrow(ws)
    for row in rows: #inputs the data made in Lon into the Relavant sheet
        for i in range(len(row)):
            ws.cell(temp, i+1, row[i])
        temp += 1
    
def Count():
    ws = wb.worksheets[2] # inputs the on time and total variables
    ws.cell(int(month)+1,2,ontim3) 
    ws.cell(int(month)+1,3,total)

def Overview():
    ws = wb.worksheets[3]
    ws.cell(int(month)+1,2,ontim3 / total) #inputs the percent

def ontime(ship, train): # figures out if the ship date is before training and install date
    global total, ontim3, month
    total = total + 1
    ship = ship.split("-")
    train = train.split("-")
    month = train[0]
    if ship[0] <= train[0] and ship[1] <= train[1] and ship[2] <= train[2]:
        ontim3 = ontim3 + 1
        return("Yes")
    else: return("No")

def getmaxrow(ws):
    maybe = 0
    for i in range(1, ws.max_row):
        if (ws.cell(i,1).value == "" or ws.cell(i,1).value == None) and maybe == 1:
            return(i)
        elif ws.cell(i,1).value == "" or ws.cell(i,1).value == None:
            maybe = 1
        else:
            maybe = 0

def main():
    Eval()
    Dates()
    Count()
    Overview()
    print("QO3 Done !")
    wb.save(excel_file)

main()
