import csv
import sys
from openpyxl import load_workbook

#These 2 are the only variables you need to change.(normally you can just pass them in through command line arguments i.e. "python QO2.py $excelfile $csv1 $csv2")
excel_file = sys.argv[1]
purchasecsv = sys.argv[2]
warantycsv = sys.argv[3]

#no touchy the rest unless it is borked
#if borked the function names are what sheet in the excel file it changes

def purchase():
    #load Purchases sheet
    ws = wb.worksheets[1]
    wsname = wb.sheetnames[1]

    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 1)
    ws = wb.worksheets[1]

    #import the CSV data
    try:
        with open(purchasecsv) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
    except FileNotFoundError as fnf_error:
        print(fnf_error)

def waranty():
    #load Waranty sheet
    ws = wb.worksheets[2]
    wsname = wb.sheetnames[2]

    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 2)
    ws = wb.worksheets[2]

    #import the CSV data
    try:
        with open(warantycsv) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
    except FileNotFoundError as fnf_error:
       print(fnf_error)
       
def overlap():
    #load Waranty sheet
    ws = wb.worksheets[3]
    wsname = wb.sheetnames[3]

    #Clear the sheet
    wb.remove(ws)
    wb.create_sheet(wsname, 3)
    ws = wb.worksheets[3]
    
    # load data from worksheets
    wsp = wb.worksheets[1]
    wsw = wb.worksheets[2]
    cmax = wsw.max_row

    #sanity variables
    global punum
    global returns
    global date
    punum = (int(wsp.max_row) - 1)
    pssn,wssn,returns,data = [],[],[],[]
    date = wsw.cell(2,1).value[1:-6]

    #transfer and transform
    #Creates rows for the overlap sheet
    for i in range(1, cmax+1):
        if i != 1:
            pssn.append(wsp.cell(i,4).value)
            wssn.append(wsw.cell(i,4).value)
        data.append([wsp.cell(i,1).value,wsp.cell(i,2).value,wsp.cell(i,4).value,wsp.cell(i,5).value,None,None,wsw.cell(i,1).value,wsw.cell(i,2).value,wsw.cell(i,4).value,wsw.cell(i,5).value,wsw.cell(i,6).value])
   
    #finds the returns and adds them to the "returns" array
    for i in range(len(pssn)):
        for j in range(len(pssn)):
            if bool(pssn[i] == wssn[j]) & bool(pssn[i] != None) & bool(pssn[i] != ''):
                returns.append(wsp.cell(i+2,2).value)
   
    #adds the rows to the overlap sheet
    for row in data:
        ws.append(row)

def ret():
    #simply adds the precalculated returns 
     ws = wb.worksheets[4]
     for i in range(1, len(returns)+1):
         ws.cell(i+1,int(date),returns[i-1])

def over():
    #updates the oversite sheet 
     ws = wb.worksheets[0]
     ws.cell(int(date)+1,3,punum)
     ws.cell(int(date)+1,4,len(returns))

def main():
    global wb
    wb = load_workbook(excel_file)
    purchase()
    waranty()
    overlap()
    ret()
    over()
    print("QO2 Done!")
    wb.save(excel_file)

main()