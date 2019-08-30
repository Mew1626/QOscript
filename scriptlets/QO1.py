import sys
from openpyxl import load_workbook

excel_file = sys.argv[1]
reports = int(sys.argv[2])
under10min = int(sys.argv[3])
month = int(sys.argv[4])

def main():
    wb = load_workbook(excel_file)
    ws = wb.worksheets[0]
    ws.cell(month,3,reports)
    ws.cell(month,4,under10min)
    print("QO1 Done!")
    wb.save(excel_file)

main()